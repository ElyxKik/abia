#!/usr/bin/env python3
"""
Module de traitement des fichiers Excel pour l'intégration avec l'API DeepSeek.
Ce script extrait les données d'un fichier Excel et les formate pour l'envoi à l'API DeepSeek.
"""

import sys
import os
import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import xlrd
import traceback

def convert_sections_to_array(sections_dict):
    """
    Convertit un dictionnaire de sections en tableau pour compatibilité avec le visualiseur JavaScript
    
    Args:
        sections_dict (dict): Dictionnaire de sections avec clé=titre et valeur=contenu
        
    Returns:
        list: Tableau de sections au format attendu par le visualiseur
    """
    sections_array = []
    for section_name, section_content in sections_dict.items():
        sections_array.append({
            "title": section_name,
            "content": section_content
        })
    return sections_array

def process_excel_file(file_path, format_type='markdown', max_rows=100, max_cols=20, sheet_index=0):
    """
    Traite un fichier Excel et extrait les données dans le format spécifié.
    
    Args:
        file_path (str): Chemin vers le fichier Excel
        format_type (str): Format de sortie ('json', 'markdown', 'csv', 'text')
        max_rows (int): Nombre maximum de lignes à extraire
        max_cols (int): Nombre maximum de colonnes à extraire
        sheet_index (int): Index de la feuille à traiter (0 = première feuille)
        
    Returns:
        dict: Métadonnées du fichier et données extraites
    """
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return {"error": f"Le fichier {file_path} n'existe pas"}
        
        # Charger le fichier Excel avec pandas
        if file_ext in ('.xlsx', '.xls', '.xlsm', '.csv'):
            try:
                # Essayer d'abord d'ouvrir le fichier comme CSV si l'extension est .csv
                if file_ext == '.csv':
                    try:
                        df = pd.read_csv(file_path, nrows=max_rows)
                        sheet_names = ["CSV Data"]
                        sheet_count = 1
                        sheet_index = 0
                    except Exception as csv_error:
                        return {"error": f"Erreur lors de la lecture du fichier CSV: {str(csv_error)}"}
                else:
                    # Obtenir le nombre de feuilles et leurs noms
                    sheet_names = []
                    try:
                        if file_ext == '.xlsx' or file_ext == '.xlsm':
                            try:
                                workbook = load_workbook(file_path, read_only=True, data_only=True)
                                sheet_names = workbook.sheetnames
                                sheet_count = len(sheet_names)
                            except Exception as openpyxl_error:
                                if "not a zip file" in str(openpyxl_error).lower() or "central directory" in str(openpyxl_error).lower():
                                    # Essayer de lire le fichier comme CSV si le format XLSX est corrompu
                                    try:
                                        df = pd.read_csv(file_path, nrows=max_rows)
                                        sheet_names = ["Recovered Data"]
                                        sheet_count = 1
                                        sheet_index = 0
                                        print(f"Fichier Excel corrompu, lu comme CSV: {file_path}", file=sys.stderr)
                                    except:
                                        # Si cela échoue également, réessayer avec xlrd comme dernier recours
                                        try:
                                            workbook = xlrd.open_workbook(file_path, on_demand=True, formatting_info=False)
                                            sheet_names = workbook.sheet_names()
                                            sheet_count = workbook.nsheets
                                        except Exception as xlrd_error:
                                            return {"error": f"Format de fichier Excel non reconnu ou corrompu. Essayez de le réenregistrer au format .xlsx ou .csv. Détails: {str(openpyxl_error)}"}
                                else:
                                    raise openpyxl_error
                        else:  # .xls
                            try:
                                workbook = xlrd.open_workbook(file_path, on_demand=True, formatting_info=False)
                                sheet_names = workbook.sheet_names()
                                sheet_count = workbook.nsheets
                            except Exception as xlrd_error:
                                if "not a .xls file" in str(xlrd_error).lower():
                                    # Essayer de lire le fichier comme CSV si le format XLS est corrompu
                                    try:
                                        df = pd.read_csv(file_path, nrows=max_rows)
                                        sheet_names = ["Recovered Data"]
                                        sheet_count = 1
                                        sheet_index = 0
                                        print(f"Fichier Excel corrompu, lu comme CSV: {file_path}", file=sys.stderr)
                                    except:
                                        return {"error": f"Format de fichier Excel non reconnu ou corrompu. Essayez de le réenregistrer au format .xlsx ou .csv. Détails: {str(xlrd_error)}"}
                                else:
                                    raise xlrd_error
                        
                        # Vérifier que l'index de feuille est valide
                        if sheet_index >= sheet_count:
                            sheet_index = 0
                    except Exception as sheet_error:
                        return {"error": f"Erreur lors de la lecture des feuilles du fichier Excel: {str(sheet_error)}"}
                    
                    # Si nous n'avons pas encore de DataFrame (cas où nous avons récupéré les feuilles avec succès)
                    if 'df' not in locals():
                        # Déterminer le moteur en fonction de l'extension du fichier
                        engine = None
                        if file_ext == '.xls':
                            engine = 'xlrd'
                        elif file_ext in ('.xlsx', '.xlsm'):
                            engine = 'openpyxl'
                        
                        try:
                            # Charger la feuille spécifiée avec pandas, en spécifiant le moteur
                            df = pd.read_excel(file_path, sheet_name=sheet_index, nrows=max_rows, engine=engine)
                        except ValueError as ve:
                            if "'nrows' must be an integer >=0" in str(ve) and engine == 'xlrd':
                                sheet_name_to_read = sheet_names[sheet_index] if sheet_index < len(sheet_names) else sheet_index
                                print(f"Avertissement: La lecture partielle (avec nrows) du fichier .xls '{file_path}' (feuille: {sheet_name_to_read}) a échoué avec xlrd. Tentative de lecture complète puis sélection des premières {max_rows} lignes.", file=sys.stderr)
                                try:
                                    df_full = pd.read_excel(file_path, sheet_name=sheet_index, engine=engine)
                                    df = df_full.head(max_rows)
                                    print(f"Lecture complète et sélection des {len(df)} premières lignes réussies pour '{file_path}'.", file=sys.stderr)
                                except Exception as e_fallback:
                                    print(f"Erreur lors de la tentative de lecture complète du fichier Excel '{file_path}' après l'échec de nrows: {str(e_fallback)}", file=sys.stderr)
                                    return {"error": f"Impossible de lire le contenu du fichier Excel. Le fichier pourrait être corrompu. Détails (fallback xlrd): {str(e_fallback)}"}
                            else:
                                print(f"Erreur pandas (ValueError non gérée spécifiquement ou moteur autre que xlrd): {str(ve)} pour {file_path}", file=sys.stderr)
                                return {"error": f"Erreur lors de la lecture du fichier Excel avec pandas. Détails: {str(ve)}"}
                        except Exception as pd_error:
                            print(f"Erreur pandas (générique) lors de la lecture de '{file_path}': {str(pd_error)}", file=sys.stderr)
                            # Essayer une dernière fois avec des options plus permissives
                            try:
                                if engine == 'openpyxl':
                                    # Pour les fichiers XLSX, essayer avec des options plus permissives
                                    df = pd.read_excel(file_path, sheet_name=sheet_index, nrows=max_rows, engine=engine, 
                                                      keep_default_na=True, na_values=['NA'], convert_float=True)
                                else:
                                    # Pour les autres formats (ou si le moteur n'est pas openpyxl), relancer l'erreur originale
                                    # car la logique de fallback spécifique à openpyxl ne s'applique pas.
                                    return {"error": f"Impossible de lire le contenu du fichier Excel. Le fichier pourrait être corrompu ou dans un format non supporté. Détails: {str(pd_error)}"}
                            except Exception as final_read_error:
                                return {"error": f"Impossible de lire le contenu du fichier Excel même avec des options permissives. Le fichier pourrait être sévèrement corrompu. Détails: {str(final_read_error)}"}
                
                # Limiter le nombre de colonnes
                if len(df.columns) > max_cols:
                    df = df.iloc[:, :max_cols]
                
                # Obtenir le nom de la feuille
                sheet_name = sheet_names[sheet_index] if sheet_names else f"Sheet {sheet_index+1}"
                
                # Remplacer les valeurs NaN par None pour la sérialisation JSON
                df = df.replace({np.nan: None})
                
                # Extraire les métadonnées
                file_name = os.path.basename(file_path)
                row_count = min(len(df), max_rows)
                col_count = min(len(df.columns), max_cols)
                
                # Formater les données selon le format demandé
                if format_type == 'json':
                    # Convertir en liste de dictionnaires
                    data = df.to_dict(orient='records')
                elif format_type == 'markdown':
                    # Convertir en tableau Markdown
                    data = df.to_markdown(index=False)
                elif format_type == 'csv':
                    # Convertir en CSV
                    data = df.to_csv(index=False)
                else:  # text
                    # Convertir en texte simple
                    data = df.to_string(index=False)
                
                # Retourner les métadonnées et les données
                return {
                    "fileName": file_name,
                    "sheetName": sheet_name,
                    "rowCount": row_count,
                    "columnCount": col_count,
                    "format": format_type,
                    "data": data,
                    "sheet_count": sheet_count,
                    "sheet_names": sheet_names
                }
            
            except Exception as e:
                error_msg = f"Erreur lors du traitement du fichier Excel: {str(e)}"
                traceback.print_exc()
                return {"error": error_msg}
        else:
            return {"error": f"Le format de fichier {file_ext} n'est pas pris en charge"}
    
    except Exception as e:
        error_msg = f"Erreur inattendue: {str(e)}"
        traceback.print_exc()
        return {"error": error_msg}

def generate_structured_report(data, instructions=None, llm_analysis=None):
    """
    Génère un rapport structuré au format JSON à partir des données Excel
    
    Args:
        data (dict): Données Excel extraites
        instructions (str): Instructions spécifiques pour l'analyse
        llm_analysis (str): Analyse LLM précédemment générée (optionnel)
        
    Returns:
        dict: Rapport structuré au format JSON
    """
    
    # Vérifier et initialiser les paramètres
    if llm_analysis is None:
        llm_analysis = ""
        print("Aucune analyse LLM fournie, génération d'un rapport standard", file=sys.stderr)
    
    # S'assurer que instructions est une chaîne
    if instructions is None:
        instructions = ""
    try:
        # Extraire les métadonnées du fichier
        file_info = {
            "nom": data.get("fileName", ""),
            "feuille": data.get("sheetName", ""),
            "dimensions": f"{data.get('rowCount', 0)} lignes × {data.get('columnCount', 0)} colonnes"
        }
        
        # Analyser les données pour extraire des sections pertinentes
        df = None
        sections = {}
        recommandations = []
        calculs_exemple = {}
        
        # Convertir les données en DataFrame si elles sont au format JSON
        if data.get("format") == "json" and isinstance(data.get("data"), list):
            df = pd.DataFrame(data.get("data"))
        elif data.get("format") == "markdown" and isinstance(data.get("data"), str):
            # Pour les données markdown, on essaie de les reconvertir en DataFrame
            try:
                # Créer un DataFrame à partir des données brutes pour l'analyse
                file_ext = os.path.splitext(data.get("fileName", ""))[1].lower()
                engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
                df_temp = pd.read_excel(data.get("_file_path", ""), sheet_name=data.get("_sheet_index", 0), 
                                        nrows=data.get("rowCount", 100), engine=engine)
                df = df_temp.replace({np.nan: None})
            except Exception as e:
                print(f"Erreur lors de la conversion des données markdown en DataFrame: {str(e)}", file=sys.stderr)
        
        if df is not None:
            # Identifier les colonnes numériques
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            
            # Statistiques de base
            if len(numeric_cols) > 0:
                sections["Statistiques"] = {}
                for col in numeric_cols[:5]:  # Limiter à 5 colonnes pour éviter la surcharge
                    try:
                        sections["Statistiques"][col] = {
                            "Moyenne": round(df[col].mean(), 2),
                            "Médiane": round(df[col].median(), 2),
                            "Min": round(df[col].min(), 2),
                            "Max": round(df[col].max(), 2)
                        }
                    except:
                        pass
            
            # Détecter les sections potentielles basées sur les noms de colonnes
            finance_keywords = ["montant", "prix", "coût", "revenu", "dépense", "taux", "impôt", "valeur"]
            date_keywords = ["date", "année", "mois", "jour", "période"]
            product_keywords = ["produit", "article", "référence", "stock", "quantité"]
            
            # Créer des sections basées sur les mots-clés détectés
            finance_cols = [col for col in df.columns if any(kw in str(col).lower() for kw in finance_keywords)]
            date_cols = [col for col in df.columns if any(kw in str(col).lower() for kw in date_keywords)]
            product_cols = [col for col in df.columns if any(kw in str(col).lower() for kw in product_keywords)]
            
            if finance_cols:
                sections["Finances"] = {}
                for col in finance_cols[:5]:
                    try:
                        # Prendre les 3 premières valeurs uniques comme exemples
                        unique_vals = df[col].dropna().unique()[:3]
                        sections["Finances"][col] = [float(val) if isinstance(val, (int, float)) else str(val) for val in unique_vals]
                    except:
                        sections["Finances"][col] = "Données variables"
            
            if date_cols:
                sections["Périodes"] = {}
                for col in date_cols[:3]:
                    try:
                        # Prendre les 3 premières valeurs uniques comme exemples
                        unique_vals = df[col].dropna().unique()[:3]
                        sections["Périodes"][col] = [str(val) for val in unique_vals]
                    except:
                        sections["Périodes"][col] = "Données variables"
            
            if product_cols:
                sections["Produits"] = {}
                for col in product_cols[:3]:
                    try:
                        # Prendre les 3 premières valeurs uniques comme exemples
                        unique_vals = df[col].dropna().unique()[:3]
                        sections["Produits"][col] = [str(val) for val in unique_vals]
                    except:
                        sections["Produits"][col] = "Données variables"
            
            # Générer des recommandations basées sur l'analyse des données
            if len(numeric_cols) > 0:
                recommandations.append("Vérifier les valeurs extrêmes dans les colonnes numériques")
            
            if df.isnull().sum().sum() > 0:
                recommandations.append("Traiter les valeurs manquantes dans le jeu de données")
            
            # Ajouter des exemples de calculs si des colonnes numériques sont présentes
            if len(numeric_cols) >= 2:
                calculs_exemple[f"Somme de {numeric_cols[0]}"] = f"Total: {round(df[numeric_cols[0]].sum(), 2)}"
                if len(numeric_cols) >= 2:
                    calculs_exemple[f"Rapport {numeric_cols[0]}/{numeric_cols[1]}"] = f"Formule: {numeric_cols[0]} / {numeric_cols[1]}"
        
        # Intégrer l'analyse LLM si disponible
        if llm_analysis and isinstance(llm_analysis, str) and llm_analysis.strip():
            print(f"Intégration de l'analyse LLM dans le rapport structuré", file=sys.stderr)
            
            # Déterminer si l'analyse LLM contient du JSON
            llm_sections = []
            llm_recommendations = []
            llm_title = None
            llm_summary = None
            
            try:
                # Essayer d'extraire du JSON de l'analyse LLM
                if '{' in llm_analysis and '}' in llm_analysis:
                    # Recherche de structures JSON valides dans la réponse LLM
                    import re
                    import json
                    
                    # Rechercher les blocs JSON potentiels
                    json_matches = re.findall(r'\{[^{}]*((\{[^{}]*\})[^{}]*)*\}', llm_analysis)
                    for potential_json, _ in json_matches:
                        try:
                            json_data = json.loads(potential_json)
                            if isinstance(json_data, dict):
                                if 'title' in json_data or 'titre' in json_data:
                                    llm_title = json_data.get('title') or json_data.get('titre')
                                
                                if 'summary' in json_data or 'resume' in json_data:
                                    llm_summary = json_data.get('summary') or json_data.get('resume')
                                
                                # Extraction des sections 
                                llm_json_sections = json_data.get('sections', [])
                                if isinstance(llm_json_sections, list):
                                    for section in llm_json_sections:
                                        if isinstance(section, dict):
                                            title = section.get('title') or section.get('titre') or 'Section sans titre'
                                            content = section.get('content') or section.get('contenu') or ''
                                            llm_sections.append({'titre': title, 'contenu': content})
                                elif isinstance(llm_json_sections, dict):
                                    for title, content in llm_json_sections.items():
                                        llm_sections.append({'titre': title, 'contenu': content if isinstance(content, str) else str(content)})
                                
                                # Extraction des recommandations
                                recommendations = json_data.get('recommendations') or json_data.get('recommandations') or []
                                if isinstance(recommendations, list):
                                    for rec in recommendations:
                                        if isinstance(rec, str):
                                            llm_recommendations.append(rec)
                                        elif isinstance(rec, dict) and ('text' in rec or 'texte' in rec):
                                            llm_recommendations.append(rec.get('text') or rec.get('texte'))
                            break  # Utiliser le premier JSON valide trouvé
                        except json.JSONDecodeError:
                            continue
                
                # Si aucun JSON valide n'a été trouvé, extraire les sections par analyse de texte
                if not llm_sections:
                    # Détecter les sections par titres
                    import re
                    section_pattern = r'\n\s*(\d+\.\s*|#{1,3}\s*|\*\*\s*)?([A-Z][^\n\.:]+)(?:\:|\n)'  
                    section_matches = re.findall(section_pattern, '\n' + llm_analysis)
                    
                    if section_matches:
                        prev_end = 0
                        for i, (prefix, title) in enumerate(section_matches):
                            title = title.strip()
                            # Déterminer le début et la fin de la section actuelle
                            start_pos = llm_analysis.find(title, prev_end)
                            if start_pos != -1:
                                end_pos = len(llm_analysis)
                                if i < len(section_matches) - 1:
                                    next_title = section_matches[i+1][1].strip()
                                    next_pos = llm_analysis.find(next_title, start_pos + len(title))
                                    if next_pos != -1:
                                        end_pos = next_pos
                                
                                # Extraire le contenu de la section
                                section_text = llm_analysis[start_pos + len(title):end_pos].strip()
                                # Nettoyer le début du texte qui peut contenir des caractères de formatage
                                section_text = section_text.lstrip(':\n- *')
                                
                                llm_sections.append({
                                    'titre': title,
                                    'contenu': section_text
                                })
                                
                                prev_end = end_pos
                    
                    # Si aucune section n'a été détectée, créer une section générale
                    if not llm_sections:
                        llm_sections.append({
                            'titre': 'Analyse générale',
                            'contenu': llm_analysis
                        })
                
                # Extraire des recommandations si pas encore fait
                if not llm_recommendations:
                    recommendation_patterns = [
                        r'(?:Recommandation|Recommendation)[s]?\s*:?\s*\n+((?:\s*-\s*[^\n]+\n*)+)',
                        r'(?:Conseil|Conseils|Suggestion|Suggestions)\s*:?\s*\n+((?:\s*-\s*[^\n]+\n*)+)',
                        r'(?:Je recommande|Je suggère)\s*:?\s*\n+((?:\s*-\s*[^\n]+\n*)+)'
                    ]
                    
                    for pattern in recommendation_patterns:
                        matches = re.search(pattern, llm_analysis, re.IGNORECASE)
                        if matches:
                            rec_text = matches.group(1)
                            rec_items = re.findall(r'\s*-\s*([^\n]+)', rec_text)
                            llm_recommendations.extend(rec_items)
                            break
            except Exception as e:
                print(f"Erreur lors de l'analyse de la réponse LLM: {str(e)}", file=sys.stderr)
            
            # Intégrer les résultats de l'analyse LLM dans le rapport final
            report_sections = {}
            
            # Combiner les sections détectées automatiquement avec celles du LLM
            if llm_sections:
                # Utiliser les sections du LLM comme base et ajouter celles détectées automatiquement
                for section in llm_sections:
                    section_title = section['titre']
                    section_content = section['contenu']
                    report_sections[section_title] = section_content
            
            # Ajouter les sections détectées automatiquement si elles ne sont pas déjà présentes
            for section_title, section_content in sections.items():
                if section_title not in report_sections:
                    report_sections[section_title] = section_content
            
            # Combiner les recommandations
            combined_recommendations = llm_recommendations + [rec for rec in recommandations if rec not in llm_recommendations]
    except Exception as e:
        print(f"Erreur lors de l'analyse de la réponse LLM: {str(e)}", file=sys.stderr)
    
    # Générer le rapport final
    try:
        # Intégrer les résultats de l'analyse LLM dans le rapport final
        report_sections = {}

        # Combiner les sections détectées automatiquement avec celles du LLM
        if llm_sections:
            # Utiliser les sections du LLM comme base et ajouter celles détectées automatiquement
            for section in llm_sections:
                section_title = section['titre']
                section_content = section['contenu']
                report_sections[section_title] = section_content

            # Ajouter les sections détectées automatiquement si elles ne sont pas déjà présentes
            for section_title, section_content in sections.items():
                if section_title not in report_sections:
                    report_sections[section_title] = section_content

            # Combiner les recommandations
            combined_recommendations = llm_recommendations + [rec for rec in recommandations if rec not in llm_recommendations]

            # Créer le rapport structuré enrichi par le LLM
            sections_array = []
            for section_name, section_content in report_sections.items():
                sections_array.append({
                    "title": section_name,
                    "content": section_content
                })

            report = {
                "titre": llm_title or "Rapport d'analyse Excel",
                "fichier": file_info,
                "resume": llm_summary or "Analyse détaillée du fichier Excel avec extraction des points clés",
                "sections": sections_array,  # Utiliser le tableau au lieu du dictionnaire
                "recommandations": combined_recommendations,
                "calculs_exemple": calculs_exemple
            }
        else:
            # Créer le rapport structuré standard (sans LLM)
            sections_array = []
            for section_name, section_content in sections.items():
                sections_array.append({
                    "title": section_name,
                    "content": section_content
                })

            report = {
                "titre": "Rapport d'analyse Excel",
                "fichier": file_info,
                "resume": "Analyse de base des données Excel",
                "sections": sections_array,  # Utiliser le tableau au lieu du dictionnaire
                "recommandations": recommandations,
                "calculs_exemple": calculs_exemple
            }

        return report

    except Exception as e:
        print(f"Erreur lors de la génération du rapport structuré: {str(e)}", file=sys.stderr)
        traceback.print_exc()
        return {
            "fichier": {
                "nom": data.get("fileName", ""),
                "feuille": data.get("sheetName", ""),
                "dimensions": f"{data.get('rowCount', 0)} lignes × {data.get('columnCount', 0)} colonnes"
            },
            "sections": [],
            "recommandations": ["Vérifier le format des données", "Analyser manuellement le contenu du fichier"],
            "calculs_exemple": {}
        }

def analyze_excel_data(data, column_types=True, stats=True, preview_rows=5):
    """
    Analyse les données Excel et génère des statistiques et informations descriptives.
    
    Args:
        data (dict): Données extraites d'un fichier Excel
        column_types (bool): Inclure les types de colonnes dans l'analyse
        stats (bool): Inclure des statistiques dans l'analyse
        preview_rows (int): Nombre de lignes à inclure dans l'aperçu
        
    Returns:
        dict: Analyse des données
    """
    try:
        if "error" in data:
            return data
        
        # Récupérer les données au format JSON
        if data["format"] != "json":
            return {"error": "L'analyse nécessite des données au format JSON"}
        
        # Convertir les données en DataFrame
        df = pd.DataFrame(data["data"])
        
        analysis = {
            "fileName": data["fileName"],
            "sheetName": data["sheetName"],
            "rowCount": data["rowCount"],
            "columnCount": data["columnCount"],
            "columns": {}
        }
        
        # Ajouter l'aperçu des données
        if preview_rows > 0:
            analysis["preview"] = df.head(preview_rows).to_dict(orient='records')
        
        # Analyser chaque colonne
        for col in df.columns:
            col_analysis = {}
            
            # Détecter le type de données
            if column_types:
                # Ignorer les valeurs None pour la détection de type
                non_null_values = df[col].dropna()
                
                if len(non_null_values) == 0:
                    col_type = "unknown"
                elif pd.api.types.is_numeric_dtype(non_null_values):
                    if all(non_null_values.apply(lambda x: isinstance(x, int) or x.is_integer())):
                        col_type = "integer"
                    else:
                        col_type = "float"
                elif pd.api.types.is_datetime64_any_dtype(non_null_values):
                    col_type = "datetime"
                elif pd.api.types.is_bool_dtype(non_null_values):
                    col_type = "boolean"
                else:
                    col_type = "string"
                
                col_analysis["type"] = col_type
            
            # Calculer les statistiques
            if stats:
                col_analysis["null_count"] = df[col].isna().sum()
                col_analysis["unique_values"] = df[col].nunique()
                
                # Statistiques pour les colonnes numériques
                if pd.api.types.is_numeric_dtype(df[col]):
                    col_analysis["min"] = df[col].min()
                    col_analysis["max"] = df[col].max()
                    col_analysis["mean"] = df[col].mean()
                    col_analysis["median"] = df[col].median()
                    col_analysis["std"] = df[col].std()
                
                # Valeurs les plus fréquentes pour les colonnes textuelles
                elif col_type == "string":
                    value_counts = df[col].value_counts().head(5).to_dict()
                    col_analysis["most_common_values"] = value_counts
            
            analysis["columns"][col] = col_analysis
        
        return analysis
    
    except Exception as e:
        error_msg = f"Erreur lors de l'analyse des données: {str(e)}"
        traceback.print_exc()
        return {"error": error_msg}

def format_prompt_for_deepseek(data, instructions, include_analysis=True, max_data_length=8000):
    """
    Formate les données Excel et les instructions pour l'envoi à l'API DeepSeek.
    
    Args:
        data (dict): Données extraites d'un fichier Excel
        instructions (str): Instructions pour l'analyse
        include_analysis (bool): Inclure l'analyse automatique dans le prompt
        max_data_length (int): Longueur maximale des données à inclure dans le prompt
        
    Returns:
        str: Prompt formaté pour DeepSeek
    """
    try:
        if "error" in data:
            return f"Erreur lors du traitement du fichier Excel: {data['error']}\n\nInstructions: {instructions}"
        
        # Créer l'en-tête du prompt
        prompt = f"""
Voici des données extraites d'un fichier Excel:
Nom du fichier: {data['fileName']}
Feuille: {data['sheetName']}
Nombre de lignes: {data['rowCount']}
Nombre de colonnes: {data['columnCount']}
Format des données: {data['format']}

"""
        
        # Ajouter les instructions
        prompt += f"INSTRUCTIONS:\n{instructions}\n\n"
        
        # Ajouter l'analyse si demandée
        if include_analysis and data["format"] == "json":
            analysis = analyze_excel_data(data)
            if "error" not in analysis:
                prompt += "ANALYSE AUTOMATIQUE DES DONNÉES:\n"
                
                # Ajouter les types de colonnes
                prompt += "Types de colonnes:\n"
                for col, info in analysis["columns"].items():
                    prompt += f"- {col}: {info.get('type', 'inconnu')}\n"
                
                # Ajouter des statistiques pour les colonnes numériques
                num_cols = [col for col, info in analysis["columns"].items() 
                           if info.get("type") in ("integer", "float")]
                
                if num_cols:
                    prompt += "\nStatistiques des colonnes numériques:\n"
                    for col in num_cols:
                        info = analysis["columns"][col]
                        prompt += f"- {col}: min={info.get('min')}, max={info.get('max')}, moyenne={info.get('mean')}, médiane={info.get('median')}\n"
                
                prompt += "\n"
        
        # Ajouter les données
        prompt += "DONNÉES:\n"
        
        # Convertir les données en texte et les tronquer si nécessaire
        data_str = str(data["data"])
        if len(data_str) > max_data_length:
            data_str = data_str[:max_data_length] + "...[données tronquées pour respecter la limite de contexte]"
        
        prompt += data_str + "\n\n"
        
        # Ajouter des instructions finales
        prompt += "Sur la base de ces données et des instructions fournies, veuillez réaliser une analyse complète."
        
        return prompt
    
    except Exception as e:
        error_msg = f"Erreur lors de la création du prompt: {str(e)}"
        traceback.print_exc()
        return f"Erreur: {error_msg}\n\nInstructions: {instructions}"

def main():
    """
    Point d'entrée principal du script lorsqu'il est exécuté directement.
    Attend un chemin de fichier Excel et des options en arguments.
    """
    if len(sys.argv) < 6:
        print(json.dumps({"error": "Arguments insuffisants. Usage: python excel_processor.py <file_path> <format> <max_rows> <max_cols> <sheet_index> [instructions]"}));
        sys.exit(1)
        
    file_path = sys.argv[1]
    format_type = sys.argv[2]
    max_rows = int(sys.argv[3])
    max_cols = int(sys.argv[4])
    sheet_index = int(sys.argv[5])
    
    try:
        # Traiter le fichier Excel
        # Respecter l'ordre des paramètres défini dans la signature de la fonction
        result = process_excel_file(file_path, format_type, max_rows, max_cols, sheet_index)
        
        # Initialiser les variables pour les arguments complémentaires
        generate_report = False
        instructions = ""
        llm_analysis = None  # Initialiser explicitement
        
        # Traiter les arguments complémentaires
        i = 6
        while i < len(sys.argv):
            arg = sys.argv[i]
            
            if arg == "--structured-report":
                generate_report = True
                i += 1
                if i < len(sys.argv) and not sys.argv[i].startswith("--"):
                    instructions = sys.argv[i]
                    i += 1
            elif arg == "--llm-analysis":
                i += 1
                if i < len(sys.argv):
                    llm_analysis_file = sys.argv[i]
                    # Lire l'analyse LLM depuis le fichier temporaire
                    try:
                        if os.path.exists(llm_analysis_file):
                            with open(llm_analysis_file, 'r') as f:
                                llm_analysis = f.read()
                                print(f"Analyse LLM lue depuis le fichier temporaire: {llm_analysis_file}", file=sys.stderr)
                        else:
                            print(f"Fichier d'analyse LLM introuvable: {llm_analysis_file}", file=sys.stderr)
                    except Exception as e:
                        print(f"Erreur lors de la lecture du fichier d'analyse LLM: {str(e)}", file=sys.stderr)
                    i += 1
            elif not arg.startswith("--") and instructions == "":
                # Pour compatibilité avec l'ancien format
                instructions = arg
                i += 1
            else:
                i += 1
        
        if generate_report:
            # Générer un rapport structuré au format JSON
            print(f"Génération d'un rapport structuré {'avec analyse LLM' if llm_analysis else 'sans analyse LLM'}", file=sys.stderr)
            try:
                # Passer explicitement les arguments pour éviter les erreurs de référence
                structured_report = generate_structured_report(
                    data=result, 
                    instructions=instructions if instructions else "", 
                    llm_analysis=llm_analysis if llm_analysis else None
                )
                result["structured_report"] = structured_report
            except Exception as e:
                print(f"Erreur lors de la génération du rapport structuré: {str(e)}", file=sys.stderr)
                # Générer un rapport basique sans analyse LLM en cas d'erreur
                structured_report = generate_structured_report(result, instructions, None)
                result["structured_report"] = structured_report
        elif instructions:  # Si des instructions sont fournies, formater le prompt pour DeepSeek
            prompt = format_prompt_for_deepseek(result, instructions)
            result["prompt"] = prompt
        
        # Supprimer les champs internes avant de retourner le résultat
        result.pop("_file_path", None)
        result.pop("_sheet_index", None)

        # Imprimer le résultat au format JSON sur stdout
        print(json.dumps(result))

    except Exception as e:
        # Capturer toute exception non gérée dans main() et la retourner comme JSON
        print(json.dumps({"error": f"Erreur inattendue dans l'exécution du script Python: {str(e)}"}))
        # Optionnel: logguer l'erreur complète sur stderr pour le débogage côté serveur
        import traceback
        print(f"Traceback de l'erreur inattendue: {traceback.format_exc()}", file=sys.stderr)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # Capturer toute exception non gérée dans main() et la retourner comme JSON
        print(json.dumps({"error": f"Erreur inattendue dans l'exécution du script Python: {str(e)}"}))
        # Optionnel: logguer l'erreur complète sur stderr pour le débogage côté serveur
        import traceback
        print(f"Traceback de l'erreur inattendue: {traceback.format_exc()}", file=sys.stderr)
